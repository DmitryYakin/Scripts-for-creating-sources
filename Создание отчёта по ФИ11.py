import pandas as pd
import json
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

def setup_template(ws):
    """Настраивает шаблон таблицы с многоуровневыми заголовками"""
    thin_border = Border(left=Side(style='thin'), 
                       right=Side(style='thin'), 
                       top=Side(style='thin'), 
                       bottom=Side(style='thin'))
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Основные заголовки
    ws.merge_cells('A3:A4')
    ws['A3'] = "Субъект РФ"
    
    ws.merge_cells('B3:B4')
    ws['B3'] = "Кол-во ИЭМК граждан 18+\n(На конец месяца)\n\nчисловой"
    
    ws.merge_cells('C3:G3')
    ws['C3'] = "Кол-во проанализированных ИЭМК граждан 18+\n(На конец месяца)"
    
    ws['C4'] = "ИЭМК с недостаточными данными\nдля формирования профиля риска\n\nчисловой"
    ws['D4'] = "Выявлены высокие риски\n\nчисловой"
    ws['E4'] = "Выявлены средние риски\n\nчисловой"
    ws['F4'] = "Выявлены низкие риски\n\nчисловой"
    ws['G4'] = "Кол-во пациентов, из числа тех, у кого на конец месяца выявлены высокие или средние риски и кто прошел ПМО, ДОГВН за посл. 2 года  или находится на ДН"
    
    ws.merge_cells('H3:H4')
    ws['H3'] = "Наименование МО\n(Краткое наименование юридического лица\nв соответствии с ФРМО)\n\nтекст"
    
    ws.merge_cells('I3:I4')
    ws['I3'] = "OID МО\n\nтекст"
    
    ws.merge_cells('J3:J4')
    ws['J3'] = "Наличие прикрепленного взрослого населения\n\nОтвет в формате: Да/Нет\n\nтекст"
    
    # Новый столбец для уникальных врачей
    ws.merge_cells('K3:K4')
    ws['K3'] = "Всего уникальных врачей\n(все специальности)\n\nчисловой"
    
    # Перемещаем старый столбец "Всего" на позицию K
    ws.merge_cells('L3:L4')
    ws['L3'] = "Всего\n(все специальности)"
    
    # Обновляем буквы для специальностей (сдвигаем на 1 столбец)
    specialties = [
        "Инфекционные болезни", "Кардиология", "Лечебное дело",
        "Неврология", "Общая врачебная практика (семейная медицина)",
        "Онкология", "Терапия", "Эндокринология", "Акушерское дело"
    ]
    
    for i, spec in enumerate(specialties):
        col = 13 + i*2  # Теперь начинаем с M (13) вместо L (12)
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
        ws.cell(row=3, column=col, value=spec)
        
        ws.cell(row=4, column=col, 
               value="Количество уникальных врачей,\nвоспользовавшихся сервисом\nMedicBK или WEBIOMED (если применимо)\nВ текущем месяце\n\nчисловой")
        ws.cell(row=4, column=col+1, 
               value="Количество обращений\nк сервису MedicBK или WEBIOMED\n(если применимо)\nВ текущем месяце\n\nчисловой")
    
    # Применяем стили
    for row in ws.iter_rows(min_row=3, max_row=4):
        for cell in row:
            cell.alignment = center_alignment
            cell.font = header_font
            cell.border = thin_border
    
    # Обновляем ширину столбцов с учетом нового
    column_widths = {
        'A': 15, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 12,
        'G': 15, 'H': 50, 'I': 15, 'J': 15, 'K': 15,
        'L': 15, 'M': 15, 'N': 15, 'O': 15, 'P': 15, 'Q': 15,
        'R': 15, 'S': 15, 'T': 15, 'U': 15, 'V': 15, 'W': 15,
        'X': 15, 'Y': 15, 'Z': 15, 'AA': 15, 'AB': 15, 'AC': 15, 'AD': 15
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

def load_iemk_risks_data():
    """Загружает данные ИЭМК и рисков из CSV файлов"""
    iemk_df = pd.read_csv('iemk.csv')
    iemk_data = dict(zip(pd.to_datetime(iemk_df['date']).dt.strftime('%Y-%m-%d'), iemk_df['value']))


    dn_df = pd.read_csv('dn.csv')
    dn_data = dict(zip(pd.to_datetime(dn_df['date']).dt.strftime('%Y-%m-%d'), dn_df['value']))
    
    risks_df = pd.read_csv('risks.csv')
    risks_data = {}
    
    for date in pd.to_datetime(risks_df['date']).unique():
        date_str = date.strftime('%Y-%m-%d')
        risks_data[date_str] = {
            'high': risks_df[(risks_df['date'] == date_str) & 
                           (risks_df['risk'].isin(['red', 'very_high', 'extremely']))]['value'].sum(),
            'medium': risks_df[(risks_df['date'] == date_str) & 
                             (risks_df['risk'] == 'yellow')]['value'].sum(),
            'low': risks_df[(risks_df['date'] == date_str) & 
                          (risks_df['risk'] == 'green')]['value'].sum()
        }
    
    return iemk_data, risks_data, dn_data

def process_yak_file(month_num):
    """Обрабатывает файл yak[month_num].csv и возвращает данные"""
    csv_file = f'm{month_num}.csv'
    if not os.path.exists(csv_file):
        print(f"Файл {csv_file} не найден")
        return None
    
    print(f"Обработка файла: {csv_file}")
    df = pd.read_csv(csv_file)
    
    # Обрабатываем специальности
    df['speciality'] = df['doctor_speciality'].apply(
        lambda x: json.loads(x)[0]['speciality_name'] if pd.notna(x) and x != '[]' else 'Не указана специальность'
    )
    
    # Добавляем названия МО
    if os.path.exists('mo.csv'):
        mo_df = pd.read_csv('mo.csv')
        mo_dict = dict(zip(mo_df['id'], mo_df['name']))
        df['hospital_name'] = df['hospital_id'].map(mo_dict)
        df['hospital_name'] = df['hospital_name'].fillna('МО не указано')
        df['hospital_name'] = df['hospital_name'].str.upper()
        df['hospital_oid'] = df['hospital_oid'].fillna('OID НЕ УКАЗАН')
    else:
        df['hospital_name'] = 'Мо не указано'

    # Считаем общее количество обращений по каждой МО
    total_visits = df.groupby(['hospital_name', 'hospital_oid']).size().reset_index(name='Всего (все специальности)')
    
    # Считаем общее количество уникальных врачей по каждой МО (по всем специальностям)
    total_doctors = df.groupby(['hospital_name', 'hospital_oid'])['doctor_oid'].nunique().reset_index(name='Всего уникальных врачей')
    
    # Группируем по специальностям для подсчета обращений
    grouped_visits = df.groupby(['hospital_name', 'hospital_oid', 'speciality']).size().unstack(fill_value=0).reset_index()
    
    # Группируем по специальностям для подсчета уникальных врачей
    grouped_doctors = df.groupby(['hospital_name', 'hospital_oid', 'speciality'])['doctor_oid'].nunique().unstack(fill_value=0).reset_index()
    
    # Объединяем все данные
    result = pd.merge(total_visits, total_doctors, on=['hospital_name', 'hospital_oid'], how='left')
    result = pd.merge(result, grouped_visits, on=['hospital_name', 'hospital_oid'], how='left')
    
    # Добавляем информацию о количестве уникальных врачей по специальностям
    for spec in grouped_doctors.columns:
        if spec not in ['hospital_name', 'hospital_oid']:
            result = result.merge(
                grouped_doctors[['hospital_name', 'hospital_oid', spec]].rename(columns={spec: f'{spec}_doctors'}),
                on=['hospital_name', 'hospital_oid'],
                how='left'
            )
    
    # Заполняем возможные пропуски нулями
    result = result.fillna(0)
    
    # Переименовываем колонки
    result = result.rename(columns={
        'hospital_name': 'МО',
        'hospital_oid': 'OID'
    })
    
    # Переупорядочиваем столбцы, чтобы "Всего уникальных врачей" был перед "Всего (все специальности)"
    cols = result.columns.tolist()
    total_unique_idx = cols.index('Всего уникальных врачей')
    total_visits_idx = cols.index('Всего (все специальности)')
    
    # Если "Всего уникальных врачей" идет после "Всего (все специальности)", меняем их местами
    if total_unique_idx > total_visits_idx:
        cols.insert(total_visits_idx, cols.pop(total_unique_idx))
        result = result[cols]
    
    return result

def update_monthly_report(region_name, month_date, yak_data, iemk_data, risks_data, dn_data):
    """Создает/обновляет месячный отчет"""
    month_str = pd.to_datetime(month_date).strftime('%Y-%m')
    filename = f"{region_name}_{month_str}.xlsx"
    
    # Создаем новую книгу
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет"
    setup_template(ws)
    
    # Заполняем общие данные
    ws['A5'] = region_name
        
    if month_date in iemk_data:
        ws['B5'] = iemk_data[month_date]
        total_iemk = iemk_data[month_date]
    
    if month_date in risks_data:
        risks = risks_data[month_date]
        ws['D5'] = risks['high']
        ws['E5'] = risks['medium']
        ws['F5'] = risks['low']
        ws['C5'] = total_iemk - (risks['high'] + risks['medium'] + risks['low'])

    if month_date in dn_data:
        ws['G5'] = dn_data[month_date]
        
    
    # Заполняем данные из yak файла (начиная со строки 5)
    if yak_data is not None:
        for idx, row in yak_data.iterrows():
            target_row = 5 + idx
            
            # Основные данные МО
            ws[f'H{target_row}'] = row['МО']
            ws[f'I{target_row}'] = row['OID']
            ws[f'J{target_row}'] = "Да"  # Заглушка
            
            # Новый столбец с уникальными врачами
            ws[f'K{target_row}'] = row.get('Всего уникальных врачей', 0)
            
            # Старый столбец "Всего"
            ws[f'L{target_row}'] = row.get('Всего (все специальности)', 0)
            
            # Данные по специальностям (врачи и обращения)
            specialties = {
                "Инфекционные болезни": ('M', 'N'),
                "Кардиология": ('O', 'P'),
                "Лечебное дело": ('Q', 'R'),
                "Неврология": ('S', 'T'),
                "Общая врачебная практика (семейная медицина)": ('U', 'V'),
                "Онкология": ('W', 'X'),
                "Терапия": ('Y', 'Z'),
                "Эндокринология": ('AA', 'AB'),
                "Акушерское дело": ('AC', 'AD')
            }
            
            for spec, (doc_col, visit_col) in specialties.items():
                ws[f'{visit_col}{target_row}'] = row.get(spec, 0)
                ws[f'{doc_col}{target_row}'] = row.get(f'{spec}_doctors', 0)
            
            # Применяем стили к новой строке
            for col in range(1, 31):  # От A до AC
                cell = ws.cell(row=target_row, column=col)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Сохраняем отчет
    wb.save(filename)
    print(f"Создан отчет: {filename}")

def main():
    # 1. Запрашиваем название региона
    region_name = input("Введите название региона (например, 'Калужская область'): ")
    
    # 2. Загружаем данные ИЭМК и рисков
    iemk_data, risks_data, dn_data = load_iemk_risks_data()
    
    # 3. Обрабатываем каждый месяц
    months = [
        ('2025-06-01', 6)
        # ('2025-02-01', 2),
        # ('2025-03-01', 3),
        # ('2025-04-01', 4),
        # ('2025-05-01', 5)
    ]
    
    for month_date, month_num in months:
        # Обрабатываем yak файл для этого месяца
        yak_data = process_yak_file(month_num)
        
        # Создаем соответствующий отчет
        update_monthly_report(region_name, month_date, yak_data, iemk_data, risks_data, dn_data)
    
    print("Все отчеты успешно созданы!")

if __name__ == "__main__":
    main()